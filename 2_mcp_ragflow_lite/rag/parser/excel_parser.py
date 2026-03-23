"""
Excel 解析器
"""
import logging
from typing import Optional

from rag.parser.base import BaseParser
from common.registry import parser_registry

logger = logging.getLogger(__name__)


@parser_registry.register(".xlsx")
@parser_registry.register(".xls")
class ExcelParser(BaseParser):
    """Excel 文档解析器"""

    name = "Excel Parser"
    extensions = [".xlsx", ".xls"]

    def parse(self, filename: str, binary: Optional[bytes] = None):
        """解析 Excel 文件，返回 (sections, tables)"""
        import openpyxl
        import io

        try:
            if binary:
                wb = openpyxl.load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
            else:
                wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open Excel {filename}: {e}")
            return [], []

        sections = []
        tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)

            if not rows:
                continue

            # 第一行作为表头
            header = rows[0]
            html = f"<table>\n<caption>{sheet_name}</caption>\n"
            html += "<tr>" + "".join(f"<th>{h}</th>" for h in header) + "</tr>\n"
            for row in rows[1:]:
                html += "<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>\n"
            html += "</table>"
            tables.append(html)

            # 同时提取文本形式
            text_rows = []
            for row in rows:
                text_rows.append(" | ".join(c for c in row if c.strip()))
            if text_rows:
                sections.append((f"[{sheet_name}]\n" + "\n".join(text_rows), "table"))

        wb.close()
        return sections, tables


# 向后兼容
_instance = ExcelParser()
parse = _instance.parse
